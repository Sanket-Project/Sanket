"""Export router — download forecast and signal data as XLSX or CSV."""

from __future__ import annotations

import csv
import io
from typing import Any

import structlog
from fastapi import APIRouter, Request, Response
from fastapi.responses import StreamingResponse
from sqlalchemy import select

from app.models.enums import IndustryCode
from app.routers.industry_router import ActiveIndustry, TenantId

log = structlog.get_logger(__name__)
router = APIRouter(prefix="/export", tags=["export"])


@router.get("/forecast.xlsx")
async def export_forecast_xlsx(
    request: Request,
    ctx: ActiveIndustry,
    tenant_id: TenantId,
) -> Response:
    """Download the latest forecast results as an Excel workbook."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        from fastapi.responses import JSONResponse

        return JSONResponse({"error": "openpyxl not installed"}, status_code=501)

    from app.models.forecast import ForecastResult, ForecastRun
    from app.models.product import Sku

    db = request.app.state.db
    industry = IndustryCode(ctx.code)

    async with db.session(str(tenant_id)) as session:
        run = await session.scalar(
            select(ForecastRun)
            .where(
                ForecastRun.tenant_id == tenant_id,
                ForecastRun.industry == industry,
                ForecastRun.status == "completed",
            )
            .order_by(ForecastRun.completed_at.desc())
            .limit(1)
        )

        rows: list[Any] = []
        if run:
            result = await session.execute(
                select(ForecastResult, Sku.sku_code)
                .join(Sku, ForecastResult.sku_id == Sku.id, isouter=True)
                .where(
                    ForecastResult.run_id == run.id,
                    ForecastResult.tenant_id == tenant_id,
                )
                .order_by(ForecastResult.sku_id, ForecastResult.forecast_date)
                .limit(10000)
            )
            rows = result.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Forecast"

    header_fill = PatternFill("solid", fgColor="7C3AED")
    header_font = Font(color="FFFFFF", bold=True)
    headers = ["SKU ID", "SKU Code", "Forecast Date", "P10", "P50 (Median)", "P90", "Model"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_i, (fc, sku_code) in enumerate(rows, 2):
        ws.cell(row=row_i, column=1, value=str(fc.sku_id))
        ws.cell(row=row_i, column=2, value=sku_code or "")
        ws.cell(row=row_i, column=3, value=str(fc.forecast_date))
        ws.cell(row=row_i, column=4, value=float(fc.p10))
        ws.cell(row=row_i, column=5, value=float(fc.p50))
        ws.cell(row=row_i, column=6, value=float(fc.p90))
        ws.cell(row=row_i, column=7, value=fc.model_name)

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"sanket_forecast_{ctx.code}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/signals.csv")
async def export_signals_csv(
    request: Request,
    ctx: ActiveIndustry,
    tenant_id: TenantId,
    days: int = 30,
) -> StreamingResponse:
    """Download recent trend signals as CSV."""
    from datetime import UTC, datetime, timedelta

    from app.models.trend import TrendSignal

    db = request.app.state.db
    industry = IndustryCode(ctx.code)
    cutoff = datetime.now(UTC) - timedelta(days=days)

    async with db.session(str(tenant_id)) as session:
        result = await session.execute(
            select(TrendSignal)
            .where(
                TrendSignal.industry == industry,
                TrendSignal.captured_at >= cutoff,
            )
            .order_by(TrendSignal.captured_at.desc())
            .limit(5000)
        )
        signals = result.scalars().all()

    buf = io.StringIO()
    writer = csv.DictWriter(
        buf,
        fieldnames=[
            "id",
            "source",
            "kind",
            "series_key",
            "industry",
            "normalized_score",
            "confidence",
            "raw_value",
            "category_tags",
            "region",
            "captured_at",
        ],
    )
    writer.writeheader()
    for s in signals:
        writer.writerow(
            {
                "id": str(s.id),
                "source": s.source.value if hasattr(s.source, "value") else s.source,
                "kind": s.kind.value if hasattr(s.kind, "value") else s.kind,
                "series_key": s.series_key,
                "industry": s.industry.value if hasattr(s.industry, "value") else s.industry,
                "normalized_score": float(s.normalized_score),
                "confidence": float(s.confidence),
                "raw_value": float(s.raw_value) if s.raw_value is not None else "",
                "category_tags": "|".join(s.category_tags or []),
                "region": s.region or "",
                "captured_at": str(s.captured_at),
            }
        )
    buf.seek(0)

    filename = f"sanket_signals_{ctx.code}_{days}d.csv"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _csv_response(
    fieldnames: list[str], rows: list[dict[str, Any]], filename: str
) -> StreamingResponse:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(rows)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/products.csv")
async def export_products_csv(
    request: Request,
    ctx: ActiveIndustry,
    tenant_id: TenantId,
) -> StreamingResponse:
    """Download the active industry's product catalogue as CSV."""
    from app.models.product import Product

    db = request.app.state.db
    industry = IndustryCode(ctx.code)

    async with db.session(str(tenant_id)) as session:
        result = await session.execute(
            select(Product)
            .where(Product.tenant_id == tenant_id, Product.industry == industry)
            .order_by(Product.created_at.desc())
        )
        products = result.scalars().all()

    rows = [
        {
            "id": str(p.id),
            "external_id": p.external_id or "",
            "name": p.name,
            "brand": p.brand or "",
            "category": p.category,
            "subcategory": p.subcategory or "",
            "status": p.status.value if hasattr(p.status, "value") else p.status,
            "industry": p.industry.value if hasattr(p.industry, "value") else p.industry,
            "created_at": str(p.created_at),
            "updated_at": str(p.updated_at),
        }
        for p in products
    ]

    return _csv_response(
        [
            "id",
            "external_id",
            "name",
            "brand",
            "category",
            "subcategory",
            "status",
            "industry",
            "created_at",
            "updated_at",
        ],
        rows,
        f"sanket_products_{ctx.code}.csv",
    )


@router.get("/skus.csv")
async def export_skus_csv(
    request: Request,
    ctx: ActiveIndustry,
    tenant_id: TenantId,
) -> StreamingResponse:
    """Download the active industry's SKU catalogue as CSV."""
    from app.models.product import Product, Sku

    db = request.app.state.db
    industry = IndustryCode(ctx.code)

    async with db.session(str(tenant_id)) as session:
        result = await session.execute(
            select(Sku, Product.name)
            .join(Product, Product.id == Sku.product_id)
            .where(Sku.tenant_id == tenant_id, Sku.industry == industry)
            .order_by(Sku.sku_code)
        )
        records = result.all()

    rows = [
        {
            "id": str(sku.id),
            "sku_code": sku.sku_code,
            "product_name": product_name,
            "external_id": sku.external_id or "",
            "gtin": sku.gtin or "",
            "description": sku.description or "",
            "unit_cost": float(sku.unit_cost) if sku.unit_cost is not None else "",
            "unit_price": float(sku.unit_price) if sku.unit_price is not None else "",
            "currency": sku.currency,
            "lead_time_days": sku.lead_time_days if sku.lead_time_days is not None else "",
            "moq": sku.moq,
            "safety_stock": sku.safety_stock,
            "reorder_point": sku.reorder_point,
            "is_active": sku.is_active,
        }
        for sku, product_name in records
    ]

    return _csv_response(
        [
            "id",
            "sku_code",
            "product_name",
            "external_id",
            "gtin",
            "description",
            "unit_cost",
            "unit_price",
            "currency",
            "lead_time_days",
            "moq",
            "safety_stock",
            "reorder_point",
            "is_active",
        ],
        rows,
        f"sanket_skus_{ctx.code}.csv",
    )


@router.get("/inventory.csv")
async def export_inventory_csv(
    request: Request,
    ctx: ActiveIndustry,
    tenant_id: TenantId,
) -> StreamingResponse:
    """Download current warehouse stock positions as CSV."""
    from app.models.inventory import InventoryLevel
    from app.models.product import Sku

    db = request.app.state.db
    industry = IndustryCode(ctx.code)

    async with db.session(str(tenant_id)) as session:
        result = await session.execute(
            select(InventoryLevel, Sku.sku_code)
            .join(Sku, Sku.id == InventoryLevel.sku_id)
            .where(InventoryLevel.tenant_id == tenant_id, InventoryLevel.industry == industry)
            .order_by(InventoryLevel.as_of.desc())
        )
        records = result.all()

    rows = [
        {
            "sku_id": str(lvl.sku_id),
            "sku_code": sku_code,
            "location": lvl.location,
            "on_hand_units": float(lvl.on_hand_units),
            "inbound_units": float(lvl.inbound_units),
            "reserved_units": float(lvl.reserved_units),
            "available_units": float(lvl.available_units),
            "source": lvl.source,
            "as_of": str(lvl.as_of),
        }
        for lvl, sku_code in records
    ]

    return _csv_response(
        [
            "sku_id",
            "sku_code",
            "location",
            "on_hand_units",
            "inbound_units",
            "reserved_units",
            "available_units",
            "source",
            "as_of",
        ],
        rows,
        f"sanket_inventory_{ctx.code}.csv",
    )


@router.get("/alerts.csv")
async def export_alerts_csv(
    request: Request,
    ctx: ActiveIndustry,
    tenant_id: TenantId,
) -> StreamingResponse:
    """Download shortage alerts (last 30 days) as CSV."""
    from datetime import UTC, datetime, timedelta

    from app.models.alert import ShortageAlert
    from app.models.product import Sku

    db = request.app.state.db
    industry = IndustryCode(ctx.code)
    cutoff = datetime.now(UTC) - timedelta(days=30)

    async with db.session(str(tenant_id)) as session:
        result = await session.execute(
            select(ShortageAlert, Sku.sku_code)
            .join(Sku, Sku.id == ShortageAlert.sku_id, isouter=True)
            .where(
                ShortageAlert.tenant_id == tenant_id,
                ShortageAlert.industry == industry,
                ShortageAlert.fired_at >= cutoff,
            )
            .order_by(ShortageAlert.fired_at.desc())
        )
        records = result.all()

    rows = [
        {
            "id": str(a.id),
            "sku_code": sku_code or "",
            "severity": a.severity.value if hasattr(a.severity, "value") else a.severity,
            "status": a.status.value if hasattr(a.status, "value") else a.status,
            "risk_score": float(a.risk_score),
            "coverage_days": float(a.coverage_days) if a.coverage_days is not None else "",
            "p10_demand": float(a.p10_demand) if a.p10_demand is not None else "",
            "p50_demand": float(a.p50_demand) if a.p50_demand is not None else "",
            "p90_demand": float(a.p90_demand) if a.p90_demand is not None else "",
            "trend_score": float(a.trend_score) if a.trend_score is not None else "",
            "title": a.title,
            "message": a.message,
            "fired_at": str(a.fired_at),
            "acknowledged_at": str(a.acknowledged_at) if a.acknowledged_at else "",
            "resolved_at": str(a.resolved_at) if a.resolved_at else "",
        }
        for a, sku_code in records
    ]

    return _csv_response(
        [
            "id",
            "sku_code",
            "severity",
            "status",
            "risk_score",
            "coverage_days",
            "p10_demand",
            "p50_demand",
            "p90_demand",
            "trend_score",
            "title",
            "message",
            "fired_at",
            "acknowledged_at",
            "resolved_at",
        ],
        rows,
        f"sanket_alerts_{ctx.code}.csv",
    )
