"""File upload → canonical schema import (CSV / Excel).

This is the universal, always-available connector: whatever a customer's stack
is, they can export sales / inventory / products to a spreadsheet and land it in
SANKET's canonical tables (products / skus / inventory_levels / historical_sales).

Design mirrors the Shopify sync: the pure ``map_*_row`` helpers do all the field
mapping and column-alias resolution with no DB dependency (so they're unit
tested), and a thin DB orchestrator upserts the mapped rows in one tenant-scoped
transaction.

Column matching is forgiving — headers are lower-cased, stripped of spaces /
underscores, and matched against an alias set — so "Selling Price", "unit_price"
and "price" all resolve to the same field.
"""

from __future__ import annotations

import csv
import io
import uuid
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

import structlog
from sqlalchemy import select
from sqlalchemy.dialects.postgresql import insert as pg_insert

from app.models.enums import IndustryCode, ProductStatus
from app.models.inventory import InventoryLevel
from app.models.product import Product, Sku
from app.models.sales import HistoricalSale

log = structlog.get_logger(__name__)

SOURCE = "upload"

# Uploaded sale_time must land inside an existing range partition of
# historical_sales (static quarters cover ~2022-2027 + rolling maintenance).
# Rows outside this window are skipped with a clear error rather than failing
# the whole transaction with "no partition found".
_MIN_SALE_DATE = date(2022, 1, 1)


def _max_sale_date() -> date:
    return date(datetime.now(tz=UTC).year + 1, 12, 31)


# ── Column alias resolution ──────────────────────────────────────────────────
def _norm(header: str) -> str:
    return "".join(ch for ch in header.lower() if ch.isalnum())


# Each canonical field maps to a set of accepted (normalized) header aliases.
_ALIASES: dict[str, set[str]] = {
    "sku": {"sku", "skucode", "skuid", "item", "itemcode", "productcode", "variantsku", "code"},
    "quantity": {"quantity", "qty", "units", "unitssold", "sold", "count", "volume"},
    "price": {"sellingprice", "price", "unitprice", "saleprice", "amount", "rate"},
    "revenue": {"revenue", "grossrevenue", "totalrevenue", "sales", "total", "linetotal"},
    "discount": {"discount", "markdown", "discountamount"},
    "timestamp": {
        "timestamp",
        "date",
        "saledate",
        "orderdate",
        "datetime",
        "time",
        "soldat",
        "createdat",
    },
    "channel": {"channel", "source", "saleschannel", "store", "platform"},
    "order_id": {"orderid", "order", "ordernumber", "transactionid", "orderno", "invoice"},
    "region": {"region", "country", "state", "market"},
    # inventory
    "available": {
        "availablestock",
        "available",
        "onhand",
        "stock",
        "quantity",
        "qty",
        "onhandunits",
    },
    "reserved": {"reservedstock", "reserved", "allocated", "reservedunits"},
    "inbound": {"inboundstock", "inbound", "incoming", "onorder", "inboundunits"},
    "warehouse": {"warehouseid", "warehouse", "location", "dc", "site", "binlocation"},
    # products
    "name": {"name", "productname", "title", "product"},
    "brand": {"brand", "vendor", "manufacturer", "supplier"},
    "category": {"category", "producttype", "type", "department"},
    "description": {"description", "desc", "variant", "option"},
    "gtin": {"gtin", "barcode", "ean", "upc"},
}


def _build_header_map(headers: list[str]) -> dict[str, str]:
    """Map canonical field name → the actual header present in the file."""
    found: dict[str, str] = {}
    for raw in headers:
        n = _norm(raw)
        for field_name, aliases in _ALIASES.items():
            if n in aliases and field_name not in found:
                found[field_name] = raw
    return found


# ── File parsing ─────────────────────────────────────────────────────────────
def parse_table(filename: str, content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    """Parse an uploaded CSV/TSV/XLSX into (headers, list-of-row-dicts).

    Raises ValueError on an unsupported type, empty file, or missing header row.
    """
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return _parse_xlsx(content)
    if name.endswith((".csv", ".tsv", ".txt")):
        return _parse_csv(content, delimiter="\t" if name.endswith(".tsv") else None)
    raise ValueError("Unsupported file type — upload a .csv, .tsv or .xlsx file")


def _parse_csv(content: bytes, delimiter: str | None) -> tuple[list[str], list[dict[str, Any]]]:
    text_data = content.decode("utf-8-sig", errors="replace")
    sample = text_data[:4096]
    if delimiter is None:
        try:
            delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
        except csv.Error:
            delimiter = ","
    reader = csv.reader(io.StringIO(text_data), delimiter=delimiter)
    rows = list(reader)
    if not rows:
        raise ValueError("File is empty")
    headers = [h.strip() for h in rows[0]]
    records = [
        {
            headers[i]: (cell.strip() if isinstance(cell, str) else cell)
            for i, cell in enumerate(r)
            if i < len(headers)
        }
        for r in rows[1:]
        if any((c or "").strip() for c in r)
    ]
    return headers, records


def _parse_xlsx(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise ValueError(
            "Excel support is unavailable on this server — upload a CSV instead"
        ) from exc
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        first = next(rows_iter)
    except StopIteration as exc:
        raise ValueError("Workbook is empty") from exc
    headers = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(first)]
    records: list[dict[str, Any]] = []
    for row in rows_iter:
        if row is None or all(c is None for c in row):
            continue
        records.append({headers[i]: row[i] for i in range(len(headers)) if i < len(row)})
    wb.close()
    return headers, records


# ── Value coercion ───────────────────────────────────────────────────────────
def _to_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    s = str(value).strip().replace(",", "")
    # strip a leading currency symbol / trailing %
    s = "".join(ch for ch in s if ch.isdigit() or ch in ".-")
    if not s or s in (".", "-"):
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _to_int(value: Any) -> int | None:
    d = _to_decimal(value)
    return int(d) if d is not None else None


_DATE_FORMATS = (
    "%Y-%m-%d",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%d %H:%M:%S",
    "%d/%m/%Y",
    "%m/%d/%Y",
    "%d-%m-%Y",
    "%m/%d/%Y %H:%M",
    "%d/%m/%Y %H:%M",
)


def _to_datetime(value: Any) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=UTC)
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day, tzinfo=UTC)
    s = str(value).strip()
    try:
        return (
            datetime.fromisoformat(s.replace("Z", "+00:00")).replace(tzinfo=UTC)
            if "+" not in s and "Z" not in s.upper()
            else datetime.fromisoformat(s.replace("Z", "+00:00"))
        )
    except ValueError:
        pass
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).replace(tzinfo=UTC)
        except ValueError:
            continue
    return None


# ── Pure row mappers (unit tested) ───────────────────────────────────────────
class RowError(ValueError):
    """A row that can't be mapped; message is surfaced to the user."""


def map_sales_row(row: dict[str, Any], hmap: dict[str, str]) -> dict[str, Any]:
    """Map a raw row to a partial historical_sales dict (no tenant/sku_id).

    Requires a SKU, a positive quantity, and an in-window timestamp.
    """
    sku = str(row.get(hmap.get("sku", ""), "") or "").strip()
    if not sku:
        raise RowError("missing SKU")
    qty = _to_int(row.get(hmap.get("quantity", "")))
    if qty is None or qty <= 0:
        raise RowError(f"invalid quantity for SKU {sku}")
    when = _to_datetime(row.get(hmap.get("timestamp", "")))
    if when is None:
        raise RowError(f"missing/invalid date for SKU {sku}")
    if not (_MIN_SALE_DATE <= when.date() <= _max_sale_date()):
        raise RowError(f"date {when.date()} for SKU {sku} is outside the supported range")

    price = _to_decimal(row.get(hmap.get("price", "")))
    discount = _to_decimal(row.get(hmap.get("discount", ""))) or Decimal("0")
    revenue = _to_decimal(row.get(hmap.get("revenue", "")))
    if revenue is None and price is not None:
        revenue = price * qty
    net = (revenue - discount) if revenue is not None else None
    channel = str(row.get(hmap.get("channel", ""), "") or "upload").strip() or "upload"
    region = str(row.get(hmap.get("region", ""), "") or "").strip() or None
    order_id = str(row.get(hmap.get("order_id", ""), "") or "").strip() or None

    return {
        "_sku_code": sku,
        "sale_time": when,
        "channel": channel,
        "region": region,
        "units_sold": qty,
        "gross_revenue": revenue,
        "net_revenue": net,
        "metadata_": {"source": SOURCE, "order_id": order_id},
    }


def map_inventory_row(row: dict[str, Any], hmap: dict[str, str]) -> dict[str, Any]:
    sku = str(row.get(hmap.get("sku", ""), "") or "").strip()
    if not sku:
        raise RowError("missing SKU")
    available = _to_decimal(row.get(hmap.get("available", "")))
    if available is None:
        raise RowError(f"missing stock quantity for SKU {sku}")
    reserved = _to_decimal(row.get(hmap.get("reserved", ""))) or Decimal("0")
    inbound = _to_decimal(row.get(hmap.get("inbound", ""))) or Decimal("0")
    location = str(row.get(hmap.get("warehouse", ""), "") or "default").strip() or "default"
    return {
        "_sku_code": sku,
        "location": location,
        "on_hand_units": max(Decimal("0"), available),
        "reserved_units": max(Decimal("0"), reserved),
        "inbound_units": max(Decimal("0"), inbound),
    }


def map_product_row(row: dict[str, Any], hmap: dict[str, str]) -> dict[str, Any]:
    sku = str(row.get(hmap.get("sku", ""), "") or "").strip()
    if not sku:
        raise RowError("missing SKU")
    name = str(row.get(hmap.get("name", ""), "") or "").strip() or sku
    return {
        "_sku_code": sku,
        "name": name,
        "brand": (str(row.get(hmap.get("brand", ""), "") or "").strip() or None),
        "category": (str(row.get(hmap.get("category", ""), "") or "").strip() or "Uncategorized"),
        "description": (str(row.get(hmap.get("description", ""), "") or "").strip() or None),
        "gtin": (str(row.get(hmap.get("gtin", ""), "") or "").strip() or None),
        "unit_price": _to_decimal(row.get(hmap.get("price", ""))),
    }


# ── DB orchestration ─────────────────────────────────────────────────────────
_REQUIRED: dict[str, set[str]] = {
    "sales": {"sku", "quantity", "timestamp"},
    "inventory": {"sku", "available"},
    "products": {"sku"},
}


async def _ensure_import_product(
    session, tenant_id: uuid.UUID, industry: IndustryCode
) -> uuid.UUID:
    """Get-or-create the catch-all product that imported SKUs hang off of."""
    existing = await session.scalar(
        select(Product.id).where(
            Product.tenant_id == tenant_id,
            Product.external_id == "__imported__",
        )
    )
    if existing:
        return existing
    stmt = (
        pg_insert(Product)
        .values(
            tenant_id=tenant_id,
            industry=industry,
            external_id="__imported__",
            name="Imported items",
            category="Imported",
            status=ProductStatus.active,
            attributes={"source": SOURCE},
        )
        .on_conflict_do_update(
            index_elements=[Product.tenant_id, Product.external_id],
            index_where=Product.external_id.isnot(None),
            set_={"name": "Imported items"},
        )
        .returning(Product.id)
    )
    return (await session.execute(stmt)).scalar_one()


async def _load_sku_map(session, tenant_id: uuid.UUID) -> dict[str, uuid.UUID]:
    rows = await session.execute(select(Sku.sku_code, Sku.id).where(Sku.tenant_id == tenant_id))
    return dict(rows.all())


async def _get_or_create_sku(
    session,
    tenant_id: uuid.UUID,
    industry: IndustryCode,
    sku_code: str,
    sku_map: dict[str, uuid.UUID],
    import_product_id: uuid.UUID,
    *,
    unit_price: Decimal | None = None,
) -> uuid.UUID:
    if sku_code in sku_map:
        return sku_map[sku_code]
    stmt = (
        pg_insert(Sku)
        .values(
            tenant_id=tenant_id,
            product_id=import_product_id,
            industry=industry,
            sku_code=sku_code,
            unit_price=unit_price,
            attributes={"source": SOURCE},
            is_active=True,
        )
        .on_conflict_do_update(
            index_elements=[Sku.tenant_id, Sku.sku_code],
            set_={"is_active": True},
        )
        .returning(Sku.id)
    )
    sku_id = (await session.execute(stmt)).scalar_one()
    sku_map[sku_code] = sku_id
    return sku_id


async def import_rows(
    *,
    db,
    tenant_id: uuid.UUID,
    industry: IndustryCode,
    kind: str,
    headers: list[str],
    records: list[dict[str, Any]],
) -> dict[str, Any]:
    """Map and persist parsed rows for one ``kind`` (sales|inventory|products).

    Returns a stats dict matching the UploadResult schema fields.
    """
    if kind not in _REQUIRED:
        raise ValueError(f"Unknown import kind: {kind}")
    hmap = _build_header_map(headers)
    missing = _REQUIRED[kind] - set(hmap)
    if missing:
        pretty = ", ".join(sorted(missing))
        raise ValueError(f"File is missing required column(s) for {kind}: {pretty}")

    mapper = {"sales": map_sales_row, "inventory": map_inventory_row, "products": map_product_row}[
        kind
    ]
    mapped: list[dict[str, Any]] = []
    errors: list[str] = []
    for idx, row in enumerate(records, start=2):  # row 1 is the header
        try:
            mapped.append(mapper(row, hmap))
        except RowError as exc:
            if len(errors) < 50:
                errors.append(f"Row {idx}: {exc}")

    stats: dict[str, Any] = {
        "rows_total": len(records),
        "rows_imported": 0,
        "rows_skipped": len(records) - len(mapped),
        "products_created": 0,
        "skus_created": 0,
        "inventory_rows": 0,
        "sales_rows": 0,
        "errors": errors,
    }
    if not mapped:
        return stats

    upload_id = uuid.uuid4().hex
    async with db.session(str(tenant_id)) as session:
        import_product_id = await _ensure_import_product(session, tenant_id, industry)
        sku_map = await _load_sku_map(session, tenant_id)
        before = set(sku_map)

        if kind == "products":
            for m in mapped:
                # gtin lives in attributes (not the column) to avoid tripping the
                # partial-unique gtin index on dirty catalogs — same choice the
                # Shopify sync makes for the barcode field.
                attrs = {"source": SOURCE, "brand": m["brand"], "gtin": m["gtin"]}
                stmt = (
                    pg_insert(Sku)
                    .values(
                        tenant_id=tenant_id,
                        product_id=import_product_id,
                        industry=industry,
                        sku_code=m["_sku_code"],
                        description=m["description"],
                        unit_price=m["unit_price"],
                        attributes=attrs,
                        is_active=True,
                    )
                    .on_conflict_do_update(
                        index_elements=[Sku.tenant_id, Sku.sku_code],
                        set_={
                            "description": m["description"],
                            "unit_price": m["unit_price"],
                            "attributes": attrs,
                            "is_active": True,
                        },
                    )
                    .returning(Sku.id)
                )
                sku_map[m["_sku_code"]] = (await session.execute(stmt)).scalar_one()
                stats["rows_imported"] += 1

        elif kind == "inventory":
            for m in mapped:
                sku_id = await _get_or_create_sku(
                    session, tenant_id, industry, m["_sku_code"], sku_map, import_product_id
                )
                await session.execute(
                    pg_insert(InventoryLevel)
                    .values(
                        tenant_id=tenant_id,
                        sku_id=sku_id,
                        industry=industry,
                        location=m["location"],
                        on_hand_units=m["on_hand_units"],
                        reserved_units=m["reserved_units"],
                        inbound_units=m["inbound_units"],
                        source=SOURCE,
                        as_of=datetime.now(tz=UTC),
                    )
                    .on_conflict_do_update(
                        constraint="uq_inventory_tenant_sku_loc",
                        set_={
                            "on_hand_units": m["on_hand_units"],
                            "reserved_units": m["reserved_units"],
                            "inbound_units": m["inbound_units"],
                            "source": SOURCE,
                            "as_of": datetime.now(tz=UTC),
                        },
                    )
                )
                stats["inventory_rows"] += 1
                stats["rows_imported"] += 1

        else:  # sales
            for m in mapped:
                sku_id = await _get_or_create_sku(
                    session, tenant_id, industry, m["_sku_code"], sku_map, import_product_id
                )
                meta = {**m["metadata_"], "upload_id": upload_id}
                session.add(
                    HistoricalSale(
                        tenant_id=tenant_id,
                        sku_id=sku_id,
                        industry=industry,
                        sale_time=m["sale_time"],
                        channel=m["channel"],
                        region=m["region"],
                        units_sold=m["units_sold"],
                        gross_revenue=m["gross_revenue"],
                        net_revenue=m["net_revenue"],
                        metadata_=meta,
                    )
                )
                stats["sales_rows"] += 1
                stats["rows_imported"] += 1

        stats["skus_created"] = len(set(sku_map) - before)

    log.info(
        "file_import.done",
        tenant=str(tenant_id),
        kind=kind,
        imported=stats["rows_imported"],
        skipped=stats["rows_skipped"],
    )
    return stats
